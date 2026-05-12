# -*- coding: utf-8 -*-
import sys, io
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = load_workbook(r'C:\Users\LG\.claude\밸류체인컨설팅\output\담솥 지웰시티점_메뉴판_가안.xlsx')
ws = wb['가안']

print("=== 가안 시트 전체 ===")
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    if any(v is not None for v in row):
        print(row)
