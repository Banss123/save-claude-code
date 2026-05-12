"""
json_to_xlsx.py
사용법: python json_to_xlsx.py {매장명}
예: python json_to_xlsx.py "대흥육회"

출력: {매장명}_메뉴판.xlsx  (현안 / 가안 2개 시트)
"""

import json
import sys
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
except ImportError:
    print("openpyxl 설치 필요: pip install openpyxl")
    sys.exit(1)


# ── 색상 팔레트 ──────────────────────────────────────────────
DARK_GREEN   = "2E7D32"   # 헤더 배경
MID_GREEN    = "81C784"   # 그룹 구분선
LIGHT_GREEN  = "C8E6C9"   # 메뉴그룹명 셀
YELLOW       = "FFF9C4"   # 가안 변경/신설 행
WHITE        = "FFFFFF"
BORDER_COLOR = "A5D6A7"


def _fill(hex_color: str) -> PatternFill:
    return PatternFill("solid", fgColor=hex_color)


def _border(color: str = BORDER_COLOR, bottom_style: str = "thin") -> Border:
    s = Side(style="thin", color=color)
    b = Side(style=bottom_style, color="2E7D32" if bottom_style == "medium" else color)
    return Border(left=s, right=s, top=s, bottom=b)


def _col_a_width(menus: list) -> int:
    max_len = max((len(str(m.get("group_name", ""))) for m in menus), default=6)
    return max(max_len + 2, 8)


def _write_sheet(ws, menus: list, options: list, is_gaan: bool = False):
    center    = Alignment(horizontal="center", vertical="center")
    left      = Alignment(horizontal="left",   vertical="center", wrap_text=True)
    price_fmt = "#,##0"
    C_BASE    = 10

    # 열 너비
    col_a = _col_a_width(menus)
    for col, width in {
        "A": col_a,      "B": C_BASE * 3, "C": C_BASE,
        "D": 15,         "E": C_BASE * 3, "F": C_BASE * 2,
        "G": 3,
        "H": 8,          "I": C_BASE * 4, "J": 14,
        "K": 30,         "L": 10,
    }.items():
        ws.column_dimensions[col].width = width

    # 헤더
    hf = Font(bold=True, color=WHITE, size=10)
    for col, label in {
        "A": "메뉴그룹명", "B": "메뉴명",     "C": "가격",
        "D": "구성",       "E": "설명",       "F": "할당옵션(no)",
        "H": "번호(no)",   "I": "옵션그룹명", "J": "조건",
        "K": "옵션항목명", "L": "가격",
    }.items():
        cell           = ws[f"{col}1"]
        cell.value     = label
        cell.font      = hf
        cell.fill      = _fill(DARK_GREEN)
        cell.alignment = center
        cell.border    = _border()
    ws["G1"].fill   = _fill(DARK_GREEN)
    ws["G1"].border = _border()
    ws.freeze_panes = "A2"
    ws.sheet_format.defaultRowHeight = 25
    ws.sheet_format.customHeight = True
    ws.row_dimensions[1].height = 25

    # 메뉴 데이터
    row        = 2
    prev_group = None
    group_to_no = {opt["group_name"]: opt["no"] for opt in options}

    for menu in menus:
        cur_group = menu.get("group_name", "")

        if prev_group is not None and cur_group != prev_group:
            ws[f"G{row}"].fill = _fill(MID_GREEN)
            row += 1

        # 가안: changed/new → 노란색, unchanged → 흰색
        if is_gaan:
            ct       = menu.get("change_type", "unchanged")
            row_fill = YELLOW if ct in ("new", "modified") else WHITE
        else:
            row_fill = WHITE

        assigned = menu.get("assigned_options", [])
        no_list  = [str(g) if isinstance(g, int) else str(group_to_no.get(g, g))
                    for g in assigned]

        for col, val in {
            "A": cur_group,
            "B": menu.get("name", ""),
            "C": menu.get("discount_price") or menu.get("price", 0),
            "D": menu.get("composition", ""),
            "E": menu.get("description", ""),
            "F": ", ".join(no_list),
        }.items():
            cell           = ws[f"{col}{row}"]
            cell.value     = val
            cell.border    = _border()
            cell.alignment = center if col == "C" else left
            if col == "C":
                cell.number_format = price_fmt
            if col == "A":
                cell.fill = _fill(LIGHT_GREEN)
                cell.font = Font(bold=True, color="1B5E20", size=10)
            else:
                cell.fill = _fill(row_fill)
                if is_gaan and row_fill == YELLOW:
                    cell.font = Font(bold=True, size=10)

        ws[f"G{row}"].fill   = _fill(WHITE)
        ws[f"G{row}"].border = _border()
        prev_group = cur_group
        row       += 1

    # 옵션 데이터
    opt_row  = 2
    prev_no  = None

    for opt_idx, opt in enumerate(options):
        cur_no        = opt["no"]
        items         = opt.get("items", [])
        is_last_group = (opt_idx == len(options) - 1)

        if prev_no is not None and cur_no != prev_no:
            opt_row += 1  # 빈 행 구분

        # 가안: 신설/변경 옵션그룹 첫 행 → 노란색
        opt_changed = is_gaan and opt.get("change_type", "unchanged") in ("new", "modified")

        for item_idx, item in enumerate(items):
            is_first     = (item_idx == 0)
            is_last_item = (item_idx == len(items) - 1)
            b = _border(bottom_style="medium") if (is_last_item and not is_last_group) else _border()

            opt_fill = YELLOW if (opt_changed and is_first) else WHITE

            for col, val in {
                "H": cur_no if is_first else "",
                "I": opt["group_name"] if is_first else "",
                "J": opt.get("condition", "") if is_first else "",
                "K": item.get("name", ""),
                "L": item.get("price", 0),
            }.items():
                cell               = ws[f"{col}{opt_row}"]
                cell.value         = val
                cell.border        = b
                cell.fill          = _fill(opt_fill)
                cell.alignment     = center if col in ("H", "L") else left
                if col == "L":
                    cell.number_format = price_fmt
                if is_first:
                    cell.font = Font(bold=True, color="1B5E20", size=10)

            opt_row += 1
        prev_no = cur_no


def build_xlsx(store_name: str):
    output_dir = Path(__file__).parent / "output"

    # 현안 JSON
    hyeonan_path = output_dir / f"{store_name}_현안.json"
    if not hyeonan_path.exists():
        raise FileNotFoundError(f"파일 없음: {hyeonan_path}")
    with open(hyeonan_path, encoding="utf-8") as f:
        hyeonan = json.load(f)

    # 가안 JSON (없으면 현안만)
    gaan_path = output_dir / f"{store_name}_가안.json"
    gaan = None
    if gaan_path.exists():
        with open(gaan_path, encoding="utf-8") as f:
            gaan = json.load(f)

    wb = openpyxl.Workbook()
    ws_hyeonan = wb.active
    ws_hyeonan.title = "현안"
    _write_sheet(ws_hyeonan, hyeonan.get("menus", []), hyeonan.get("options", []), is_gaan=False)

    if gaan:
        ws_gaan = wb.create_sheet("가안")
        _write_sheet(ws_gaan, gaan.get("menus", []), gaan.get("options", []), is_gaan=True)

    xlsx_path = output_dir / f"{store_name}_메뉴판.xlsx"
    wb.save(xlsx_path)
    print(f"저장 완료: {xlsx_path}")


if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("사용법: python json_to_xlsx.py {매장명}")
        sys.exit(1)
    build_xlsx(sys.argv[1])
