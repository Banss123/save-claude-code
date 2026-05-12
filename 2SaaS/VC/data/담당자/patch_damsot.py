# -*- coding: utf-8 -*-
"""담솥 지웰시티점 메뉴판 가안 수정 스크립트.

수정 사항:
1. [숙성] 스테이크솥밥 → [스테이크 1.5배] 스테이크솥밥
   - 설명에서 '숙성' 및 '고기 1.5배 추가 가능' 제거
   - 할당옵션에서 '2' 제거
2. 옵션 no.2 (스테이크고기1.5배) 삭제, 이후 번호 재정렬
3. 매운가지치즈솥밥 설명에서 '· 치즈 추가로 부드럽게' 제거
"""

import sys
import io
import shutil
from pathlib import Path
import openpyxl
from openpyxl import load_workbook

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

SRC = Path(r'C:\Users\LG\.claude\밸류체인컨설팅\output\담솥 지웰시티점_메뉴판.xlsx')
DST = Path(r'C:\Users\LG\.claude\밸류체인컨설팅\output\담솥 지웰시티점_메뉴판_가안.xlsx')

# 원본 복사
shutil.copy2(SRC, DST)
wb = load_workbook(DST)
ws = wb['가안']

# ─── 열 인덱스 (1-based) ───
COL_NAME   = 2   # B: 메뉴명
COL_COMP   = 4   # D: 구성
COL_DESC   = 5   # E: 설명
COL_OPT    = 6   # F: 할당옵션(no)
COL_NO     = 8   # H: 번호(no)
COL_GRPNM  = 9   # I: 옵션그룹명
COL_COND   = 10  # J: 조건
COL_ITEM   = 11  # K: 옵션항목명
COL_PRICE  = 12  # L: 가격

max_row = ws.max_row

# ─── 1. 메뉴 행 수정 ───
for row in range(1, max_row + 1):
    name = ws.cell(row=row, column=COL_NAME).value

    # 스테이크솥밥
    if name == '[숙성] 스테이크솥밥':
        print(f"  행{row}: 스테이크솥밥 수정")
        ws.cell(row=row, column=COL_NAME).value = '[스테이크 1.5배] 스테이크솥밥'
        # 설명에서 '·숙성' 및 '· 고기 1.5배 추가 가능' 제거
        desc = ws.cell(row=row, column=COL_DESC).value or ''
        desc = desc.replace('·숙성 ', ' ')   # '손질·숙성 그리고' → '손질 그리고'
        desc = desc.replace(' · 고기 1.5배 추가 가능', '')
        ws.cell(row=row, column=COL_DESC).value = desc.strip() or None
        # 할당옵션 '2' 제거
        opt = ws.cell(row=row, column=COL_OPT).value
        if opt:
            parts = [p.strip() for p in str(opt).split(',') if p.strip() != '2']
            ws.cell(row=row, column=COL_OPT).value = ', '.join(parts) if parts else None

    # 매운가지치즈솥밥
    if name == '매운가지치즈솥밥':
        print(f"  행{row}: 매운가지치즈솥밥 수정")
        desc = ws.cell(row=row, column=COL_DESC).value or ''
        desc = desc.replace(' · 치즈 추가로 부드럽게', '')
        ws.cell(row=row, column=COL_DESC).value = desc.strip() or None

# ─── 2. 옵션 no.2 (스테이크고기1.5배) 삭제 ───
opt2_row = None
for row in range(1, max_row + 1):
    no_val = ws.cell(row=row, column=COL_NO).value
    grp_val = ws.cell(row=row, column=COL_GRPNM).value
    if no_val == 2 and grp_val and '스테이크' in str(grp_val):
        opt2_row = row
        print(f"  행{row}: 옵션 no.2 삭제")
        break

if opt2_row:
    # 옵션 그룹명/조건/항목/가격 셀 지우기
    for col in [COL_NO, COL_GRPNM, COL_COND, COL_ITEM, COL_PRICE]:
        ws.cell(row=opt2_row, column=col).value = None

# ─── 3. 옵션 번호 재정렬 (3→2, 4→3, 5→4, 6→5) ───
print("  옵션 번호 재정렬...")
for row in range(1, max_row + 1):
    no_val = ws.cell(row=row, column=COL_NO).value
    if isinstance(no_val, int) and no_val >= 3:
        ws.cell(row=row, column=COL_NO).value = no_val - 1

# ─── 4. 할당옵션 참조 번호 업데이트 (3→2, 4→3, 5→4, 6→5) ───
print("  할당옵션 참조 번호 업데이트...")
REMAP = {'3': '2', '4': '3', '5': '4', '6': '5'}
for row in range(1, max_row + 1):
    opt = ws.cell(row=row, column=COL_OPT).value
    if opt:
        parts = [p.strip() for p in str(opt).split(',')]
        new_parts = [REMAP.get(p, p) for p in parts]
        ws.cell(row=row, column=COL_OPT).value = ', '.join(new_parts)

# ─── 저장 ───
wb.save(DST)
print(f"\n완료: {DST}")
print(f"크기: {DST.stat().st_size:,} bytes")
