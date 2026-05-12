# -*- coding: utf-8 -*-
import openpyxl
import sys
import io

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\LG\.claude\밸류체인컨설팅\output\담솥 지웰시티점_메뉴판.xlsx')
print('시트:', wb.sheetnames)
for sheet in wb.sheetnames:
    ws = wb[sheet]
    print(f'\n=== {sheet} (행:{ws.max_row}, 열:{ws.max_column}) ===')
    # 헤더
    for row in ws.iter_rows(min_row=1, max_row=1, values_only=True):
        print('헤더:', row)
    # 전체 데이터
    for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        print(f'행{i}:', row)
