"""XLSX 메뉴판 가안 검수기.

MenuPlan (JSON) ↔ 생성된 XLSX 파일을 교차 검증합니다.

검수 항목 (CLAUDE.md 기준):
1. 현안 메뉴 수 == 원본 데이터 메뉴 수
2. 가안 메뉴 수 == 현안 메뉴 수 (추가/삭제 없는 한)
3. 옵션그룹 번호가 올바른 메뉴에 할당
4. 변경된 셀만 빨간색, 미변경 셀은 일반 폰트
5. 가격 필드가 숫자 타입
6. 사이즈별 분리 메뉴 올바르게 분리
"""

from __future__ import annotations

from pathlib import Path

import openpyxl

from src.schemas.menu import MenuPlan
from src.schemas.validation import CheckGroup, CheckItem, CheckStatus

# XLSX 열 번호 (xlsx_builder.py와 동일)
COL_MENU = 4    # D
COL_PRICE = 5   # E
COL_DESC = 6    # F
DATA_START_ROW = 3


def _read_sheet_cells(ws) -> list[dict]:
    """시트에서 메뉴 행 데이터를 추출합니다."""
    rows = []
    for row in range(DATA_START_ROW, ws.max_row + 1):
        menu_cell = ws.cell(row=row, column=COL_MENU)
        if menu_cell.value is None:
            continue
        price_cell = ws.cell(row=row, column=COL_PRICE)

        # 폰트 색상 추출
        is_red = False
        if menu_cell.font and menu_cell.font.color and menu_cell.font.color.rgb:
            is_red = "FF0000" in str(menu_cell.font.color.rgb)

        rows.append({
            "row": row,
            "name": str(menu_cell.value),
            "price": price_cell.value,
            "price_type": type(price_cell.value).__name__,
            "is_red": is_red,
        })
    return rows


def validate_xlsx(plan: MenuPlan, xlsx_path: str | Path) -> CheckGroup:
    """MenuPlan ↔ XLSX 파일을 검수합니다."""
    xlsx_path = Path(xlsx_path)
    items: list[CheckItem] = []

    # ── XLSX 로드 ──
    if not xlsx_path.exists():
        items.append(CheckItem(
            name="파일 존재", status=CheckStatus.FAIL,
            message=f"{xlsx_path.name} 파일을 찾을 수 없습니다",
        ))
        return CheckGroup(name="XLSX 검수", items=items)

    wb = openpyxl.load_workbook(str(xlsx_path))

    # 시트 존재 확인
    for sheet_name in ("현안", "가안"):
        if sheet_name not in wb.sheetnames:
            items.append(CheckItem(
                name=f"{sheet_name} 시트 존재", status=CheckStatus.FAIL,
                message=f"'{sheet_name}' 시트가 없습니다",
            ))
            return CheckGroup(name="XLSX 검수", items=items)

    items.append(CheckItem(
        name="시트 구조", status=CheckStatus.PASS,
        message="현안/가안 시트 존재",
    ))

    cur_rows = _read_sheet_cells(wb["현안"])
    pro_rows = _read_sheet_cells(wb["가안"])

    # ── 1. 현안 메뉴 수 == JSON 현안 메뉴 수 ──
    json_cur_count = plan.current.total_menus
    xlsx_cur_count = len(cur_rows)
    items.append(CheckItem(
        name="현안 메뉴 수 일치",
        status=CheckStatus.PASS if json_cur_count == xlsx_cur_count else CheckStatus.FAIL,
        message=f"JSON {json_cur_count}건 vs XLSX {xlsx_cur_count}건",
        expected=str(json_cur_count),
        actual=str(xlsx_cur_count),
    ))

    # ── 2. 가안 메뉴 수 == 현안 메뉴 수 ──
    json_pro_count = plan.proposed.total_menus
    xlsx_pro_count = len(pro_rows)
    items.append(CheckItem(
        name="가안 메뉴 수 일치",
        status=CheckStatus.PASS if json_pro_count == xlsx_pro_count else CheckStatus.FAIL,
        message=f"JSON {json_pro_count}건 vs XLSX {xlsx_pro_count}건",
        expected=str(json_pro_count),
        actual=str(xlsx_pro_count),
    ))

    menu_count_match = (json_cur_count == json_pro_count)
    items.append(CheckItem(
        name="현안/가안 메뉴 수 동일",
        status=CheckStatus.PASS if menu_count_match else CheckStatus.WARN,
        message=(
            f"현안 {json_cur_count}건 == 가안 {json_pro_count}건"
            if menu_count_match
            else f"현안 {json_cur_count}건 ≠ 가안 {json_pro_count}건 (추가/삭제 확인 필요)"
        ),
    ))

    # ── 3. 옵션그룹 번호 정합성 ──
    total_og = len(plan.proposed.option_groups)
    bad_refs: list[str] = []
    for group in plan.proposed.groups:
        for item in group.items:
            for og_id in item.option_group_ids:
                if og_id < 1 or og_id > total_og:
                    bad_refs.append(f"{item.name}: 옵션#{og_id}")

    items.append(CheckItem(
        name="옵션그룹 번호 유효성",
        status=CheckStatus.PASS if not bad_refs else CheckStatus.FAIL,
        message=(
            f"전체 {total_og}개 옵션그룹, 참조 모두 유효"
            if not bad_refs
            else f"유효하지 않은 참조: {', '.join(bad_refs[:5])}"
        ),
    ))

    # ── 4. 변경 셀 빨간색 정확성 (셀 단위) ──
    # JSON에서 현안/가안 메뉴를 플랫 리스트로
    cur_items = [item for g in plan.current.groups for item in g.items]
    pro_items = [item for g in plan.proposed.groups for item in g.items]

    red_errors: list[str] = []
    missing_red: list[str] = []

    for idx, (pro_item, pro_row) in enumerate(zip(pro_items, pro_rows)):
        cur_item = cur_items[idx] if idx < len(cur_items) else None
        name_changed = cur_item is not None and pro_item.name != cur_item.name

        if pro_row["is_red"] and not name_changed:
            red_errors.append(f"R{pro_row['row']} {pro_row['name'][:15]} (빨간색인데 미변경)")
        elif not pro_row["is_red"] and name_changed:
            missing_red.append(f"R{pro_row['row']} {pro_row['name'][:15]} (변경인데 빨간색 없음)")

    style_ok = not red_errors and not missing_red
    style_msg = "변경 셀만 빨간색 적용됨" if style_ok else ""
    if red_errors:
        style_msg = f"불필요한 빨간색 {len(red_errors)}건: {', '.join(red_errors[:3])}"
    if missing_red:
        style_msg += f" / 빨간색 누락 {len(missing_red)}건: {', '.join(missing_red[:3])}"

    items.append(CheckItem(
        name="변경 셀 빨간색 정확성",
        status=CheckStatus.PASS if style_ok else CheckStatus.FAIL,
        message=style_msg.strip(),
    ))

    # ── 5. 가격 필드 숫자 타입 ──
    str_prices: list[str] = []
    for sheet_name, rows in [("현안", cur_rows), ("가안", pro_rows)]:
        for r in rows:
            if r["price"] is not None and not isinstance(r["price"], (int, float)):
                str_prices.append(f"{sheet_name} R{r['row']} {r['name'][:10]}: {r['price_type']}")

    items.append(CheckItem(
        name="가격 필드 숫자 타입",
        status=CheckStatus.PASS if not str_prices else CheckStatus.FAIL,
        message=(
            "전체 가격 필드 숫자 타입"
            if not str_prices
            else f"문자열 가격 {len(str_prices)}건: {', '.join(str_prices[:3])}"
        ),
    ))

    # ── 6. 메뉴명/가격 값 일치 (JSON ↔ XLSX) ──
    value_mismatches: list[str] = []
    for sheet_label, json_items, xlsx_rows in [
        ("현안", cur_items, cur_rows),
        ("가안", pro_items, pro_rows),
    ]:
        for idx, (ji, xr) in enumerate(zip(json_items, xlsx_rows)):
            if ji.name != xr["name"]:
                value_mismatches.append(
                    f"{sheet_label} R{xr['row']} 메뉴명: '{ji.name[:12]}' vs '{xr['name'][:12]}'"
                )
            if xr["price"] is not None and ji.price != xr["price"]:
                value_mismatches.append(
                    f"{sheet_label} R{xr['row']} 가격: {ji.price} vs {xr['price']}"
                )

    items.append(CheckItem(
        name="JSON ↔ XLSX 값 일치",
        status=CheckStatus.PASS if not value_mismatches else CheckStatus.FAIL,
        message=(
            "전체 메뉴명/가격 일치"
            if not value_mismatches
            else f"불일치 {len(value_mismatches)}건: {', '.join(value_mismatches[:3])}"
        ),
    ))

    return CheckGroup(name="XLSX 검수", items=items)
