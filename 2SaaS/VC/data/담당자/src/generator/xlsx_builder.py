"""메뉴판 가안 XLSX 생성기.

MenuPlan (현안 + 가안) → XLSX 문서를 결정적으로 변환합니다.
메뉴(C~H) | I열 구분선 | 옵션(J~N) 병렬 구조.

기준: 반자동 메뉴판 가안 매뉴얼 (PDF)
"""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from src.schemas.menu import MenuPlan, MenuSheet

# ── 폰트 ──
FONT_NAME = "맑은 고딕"
FONT_SIZE = 10

# ── 색상 (PDF 매뉴얼 기준) ──
TITLE_FILL_CURRENT = PatternFill(
    start_color="1E3A32", end_color="1E3A32", fill_type="solid")
TITLE_FILL_PROPOSED = PatternFill(
    start_color="1B2D4E", end_color="1B2D4E", fill_type="solid")
TITLE_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")

HEADER_FILL = PatternFill(
    start_color="FBE4D5", end_color="FBE4D5", fill_type="solid")
HEADER_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)

GROUP_FILL = PatternFill(
    start_color="D9E8E3", end_color="D9E8E3", fill_type="solid")
GROUP_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)

OG_DATA_FILL = PatternFill(
    start_color="2B7A4B", end_color="2B7A4B", fill_type="solid")
OG_DATA_FONT = Font(
    name=FONT_NAME, size=FONT_SIZE, bold=True, color="FFFFFF")

DESC_FILL = PatternFill(
    start_color="EAF4FB", end_color="EAF4FB", fill_type="solid")
NOTE_FILL = PatternFill(
    start_color="FEF9E7", end_color="FEF9E7", fill_type="solid")

MENU_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
MENU_FONT_BOLD = Font(name=FONT_NAME, size=FONT_SIZE, bold=True)
CHANGED_FONT = Font(name=FONT_NAME, size=FONT_SIZE, bold=True, color="FF0000")
PRICE_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
OPTION_FONT = Font(name=FONT_NAME, size=FONT_SIZE)
OPTION_ID_FONT = Font(name=FONT_NAME, size=9)

THIN_BORDER = Border(
    left=Side(style="thin", color="D9D9D9"),
    right=Side(style="thin", color="D9D9D9"),
    top=Side(style="thin", color="D9D9D9"),
    bottom=Side(style="thin", color="D9D9D9"),
)

# ── 정렬 ──
CENTER_ALIGN = Alignment(horizontal="center", vertical="center",
                         wrap_text=True)
LEFT_TOP_ALIGN = Alignment(horizontal="left", vertical="top",
                           wrap_text=True)

# ── 열 너비 ──
COLUMN_WIDTHS = {
    "A": 1.0, "B": 13.0,
    "C": 16.0, "D": 26.0, "E": 10.0, "F": 28.0, "G": 13.0, "H": 14.0,
    "I": 2.0,
    "J": 4.0, "K": 22.0, "L": 38.0, "M": 7.0, "N": 9.0,
}

# ── 열 번호 ──
COL_GROUP = 3       # C
COL_MENU = 4        # D
COL_PRICE = 5       # E
COL_DESC = 6        # F
COL_NOTE = 7        # G
COL_OG_ORDER = 8    # H
COL_SEP = 9         # I
COL_OPT_NUM = 10    # J
COL_OPT_NAME = 11   # K
COL_OPT_ITEM = 12   # L
COL_OPT_ATTR = 13   # M
COL_OPT_PRICE = 14  # N

ROW_HEIGHT = 30


def _build_ref_lookup(ref: MenuSheet) -> dict[tuple[int, int], "MenuItem"]:
    """현안 시트에서 (group_idx, item_idx) → MenuItem 매핑을 생성합니다."""
    lookup: dict[tuple[int, int], object] = {}
    for gi, group in enumerate(ref.groups):
        for ii, item in enumerate(group.items):
            lookup[(gi, ii)] = item
    return lookup


def _write_sheet(ws, sheet: MenuSheet, reference: MenuSheet | None = None) -> None:
    """한 시트의 메뉴 + 옵션 데이터를 작성합니다.

    Args:
        reference: 가안 시트 작성 시 현안 데이터. 셀 단위 비교로
                   실제 변경된 값만 빨간색 적용.
    """
    # ── 열 너비 ──
    for letter, width in COLUMN_WIDTHS.items():
        ws.column_dimensions[letter].width = width

    # ── 1행: 타이틀 (C1:N1 병합) ──
    title_fill = (TITLE_FILL_CURRENT if sheet.sheet_type == "현안"
                  else TITLE_FILL_PROPOSED)
    title_cell = ws.cell(
        row=1, column=COL_GROUP,
        value=f"{sheet.store_name} \u2014 메뉴판 ({sheet.sheet_type})",
    )
    title_cell.font = TITLE_FONT
    title_cell.fill = title_fill
    title_cell.alignment = CENTER_ALIGN
    ws.merge_cells(
        start_row=1, start_column=COL_GROUP,
        end_row=1, end_column=COL_OPT_PRICE,
    )
    for col in range(COL_GROUP + 1, COL_OPT_PRICE + 1):
        ws.cell(row=1, column=col).fill = title_fill

    # ── 2행: 헤더 (전체 동일 FBE4D5) ──
    all_headers = [
        (COL_GROUP, "메뉴그룹"),
        (COL_MENU, "메뉴명"),
        (COL_PRICE, "가격(배달)"),
        (COL_DESC, "구성"),
        (COL_NOTE, "설명"),
        (COL_OG_ORDER, "옵션그룹 (적용순서)"),
        (COL_OPT_NUM, "옵션#"),
        (COL_OPT_NAME, "옵션그룹명 [조건]"),
        (COL_OPT_ITEM, "옵션 목록"),
        (COL_OPT_ATTR, "속성"),
        (COL_OPT_PRICE, "추가가격"),
    ]
    for col_num, header_text in all_headers:
        cell = ws.cell(row=2, column=col_num, value=header_text)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = CENTER_ALIGN
        cell.border = THIN_BORDER

    # ── 현안 참조 매핑 (셀 단위 비교용) ──
    ref_lookup = _build_ref_lookup(reference) if reference else {}

    # ── 메뉴 데이터 (C~H, 3행~) ──
    menu_row = 3
    for gi, group in enumerate(sheet.groups):
        for item_idx, item in enumerate(group.items):
            # 셀별 폰트 결정: 가안에서 실제 값이 변한 셀만 빨간색
            ref_item = ref_lookup.get((gi, item_idx))
            if item.is_changed and sheet.sheet_type == "가안" and ref_item:
                name_font = CHANGED_FONT if item.name != ref_item.name else MENU_FONT
                price_font = CHANGED_FONT if item.price != ref_item.price else MENU_FONT
                desc_font = CHANGED_FONT if item.description != ref_item.description else MENU_FONT
            elif item.is_changed:
                name_font = price_font = desc_font = CHANGED_FONT
            else:
                name_font = price_font = desc_font = MENU_FONT

            # C: 메뉴그룹 (첫 행만)
            c = ws.cell(row=menu_row, column=COL_GROUP)
            if item_idx == 0:
                c.value = group.name
                c.font = GROUP_FONT
                c.fill = GROUP_FILL
            c.alignment = CENTER_ALIGN
            c.border = THIN_BORDER

            # D: 메뉴명
            c = ws.cell(row=menu_row, column=COL_MENU, value=item.name)
            c.font = name_font
            c.alignment = LEFT_TOP_ALIGN
            c.border = THIN_BORDER

            # E: 가격 (숫자)
            c = ws.cell(row=menu_row, column=COL_PRICE, value=item.price)
            c.font = price_font
            c.number_format = "#,##0"
            c.alignment = CENTER_ALIGN
            c.border = THIN_BORDER

            # F: 구성 (bg)
            c = ws.cell(row=menu_row, column=COL_DESC,
                        value=item.description or None)
            c.font = desc_font
            c.fill = DESC_FILL
            c.alignment = LEFT_TOP_ALIGN
            c.border = THIN_BORDER

            # G: 설명 (bg)
            c = ws.cell(row=menu_row, column=COL_NOTE)
            c.fill = NOTE_FILL
            c.alignment = LEFT_TOP_ALIGN
            c.border = THIN_BORDER

            # H: 옵션그룹 적용순서 (9pt)
            if item.option_group_ids:
                og_text = ", ".join(str(x) for x in item.option_group_ids)
                c = ws.cell(row=menu_row, column=COL_OG_ORDER, value=og_text)
                c.font = OPTION_ID_FONT
                c.alignment = CENTER_ALIGN
                c.border = THIN_BORDER

            menu_row += 1

    # ── 옵션 데이터 (J~N, 3행~) ──
    opt_row = 3
    for og_idx, og in enumerate(sheet.option_groups):
        if og_idx > 0:
            opt_row += 1

        constraint = (f"최대{og.max_select}개" if not og.required
                      else f"필수{og.max_select}개")
        if og.max_select == 0:
            constraint = "선택" if not og.required else "필수"

        c = ws.cell(row=opt_row, column=COL_OPT_NUM, value=og_idx + 1)
        c.font = OG_DATA_FONT
        c.fill = OG_DATA_FILL
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER

        c = ws.cell(row=opt_row, column=COL_OPT_NAME,
                    value=f"{og.name}  [{constraint}]")
        c.font = OG_DATA_FONT
        c.fill = OG_DATA_FILL
        c.alignment = LEFT_TOP_ALIGN
        c.border = THIN_BORDER

        c = ws.cell(row=opt_row, column=COL_OPT_ITEM, value="옵션 목록")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER

        c = ws.cell(row=opt_row, column=COL_OPT_ATTR, value="속성")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER

        c = ws.cell(row=opt_row, column=COL_OPT_PRICE, value="추가가격")
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER_ALIGN
        c.border = THIN_BORDER

        opt_row += 1

        for oi in og.items:
            c = ws.cell(row=opt_row, column=COL_OPT_ITEM, value=oi.name)
            c.font = OPTION_FONT
            c.alignment = LEFT_TOP_ALIGN
            c.border = THIN_BORDER

            condition = "필수" if og.required else "선택"
            c = ws.cell(row=opt_row, column=COL_OPT_ATTR, value=condition)
            c.font = OPTION_FONT
            c.alignment = CENTER_ALIGN
            c.border = THIN_BORDER

            c = ws.cell(row=opt_row, column=COL_OPT_PRICE, value=oi.price)
            c.font = OPTION_FONT
            c.number_format = "#,##0"
            c.alignment = CENTER_ALIGN
            c.border = THIN_BORDER

            opt_row += 1

    # ── 행 높이 ──
    max_row = max(menu_row, opt_row, 3)
    for r in range(1, max_row):
        ws.row_dimensions[r].height = ROW_HEIGHT


def build_menu_xlsx(plan: MenuPlan, output_path: str | Path) -> Path:
    """MenuPlan → XLSX 문서를 생성합니다."""
    output_path = Path(output_path)
    wb = Workbook()

    ws_current = wb.active
    ws_current.title = "현안"
    _write_sheet(ws_current, plan.current)

    ws_proposed = wb.create_sheet(title="가안")
    _write_sheet(ws_proposed, plan.proposed, reference=plan.current)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(output_path))
    return output_path
