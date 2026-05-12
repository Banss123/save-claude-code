# -*- coding: utf-8 -*-
import sys
import io
import openpyxl

sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

wb = openpyxl.load_workbook(r'C:\Users\LG\.claude\밸류체인컨설팅\output\담솥 지웰시티점_메뉴판.xlsx')
for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    print(f'\n{"="*60}')
    print(f'시트: {sheet_name}')
    print(f'{"="*60}')
    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
        if any(v is not None for v in row):
            print(row)
