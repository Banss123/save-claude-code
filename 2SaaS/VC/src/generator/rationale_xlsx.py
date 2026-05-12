"""목표매출 산정 근거 XLSX 생성 — 담당자 내부용 (4시트).

담당자 피드백(2026-04-22): 기존 txt 버전 가독성 부족 → XLSX 4시트 구성으로 전면 교체.

시트 구성:
  1.요약       매장 메타 + 1차/2차 목표(원본/반올림 병기) + 플랫폼별 매출
  2.레버분석   4개 레버(노출·CTR·CVR·AOV) × 단기/중기 개선 여력
  3.가정·한계  산정 가정 6개 + 데이터 부족 disclaimers
  4.REF검증    정본 §필수 준수 REF-1~8 자동 검증 결과

업주에게 전달되지 않는다. orchestrator.py 에서 바탕화면 복사 대상(3번째 파일).
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from src.schemas.solution import SolutionPlan
from src.validator.reference_check import check_reference

# ─────────────────────────────────────────────
# 공용 스타일
# ─────────────────────────────────────────────
FONT_NAME = "Malgun Gothic"
BODY_FONT = Font(name=FONT_NAME, size=10)
BODY_BOLD = Font(name=FONT_NAME, size=10, bold=True)
HEADER_FONT = Font(name=FONT_NAME, size=11, bold=True, color="FFFFFF")
TITLE_FONT = Font(name=FONT_NAME, size=16, bold=True, color="FFFFFF")
SUBTITLE_FONT = Font(name=FONT_NAME, size=12, bold=True, color="FFFFFF")
META_LABEL_FONT = Font(name=FONT_NAME, size=10, bold=True)

TITLE_FILL = PatternFill(start_color="1F3864", end_color="1F3864", fill_type="solid")
SUBTITLE_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
ALT_FILL = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

# 신호등 색 (달성 확률, REF 결과)
GREEN_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
YELLOW_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
RED_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
GREY_FILL = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")

GREEN_FONT = Font(name=FONT_NAME, size=10, bold=True, color="006100")
YELLOW_FONT = Font(name=FONT_NAME, size=10, bold=True, color="9C5700")
RED_FONT = Font(name=FONT_NAME, size=10, bold=True, color="9C0006")

THIN = Side(style="thin", color="BFBFBF")
THIN_BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
LEFT = Alignment(horizontal="left", vertical="center", wrap_text=True)
RIGHT = Alignment(horizontal="right", vertical="center", wrap_text=True)

WON_FORMAT = '#,##0"원"'

_NA = "-"


# ─────────────────────────────────────────────
# 값 포맷터
# ─────────────────────────────────────────────
def _round_down_to_500k(amount: int) -> int:
    if amount < 100_000:
        return amount
    return (amount // 500_000) * 500_000


def _pct_str(value: Any, digits: int = 1) -> str:
    if value is None:
        return _NA
    try:
        return f"{float(value):.{digits}f}%"
    except (TypeError, ValueError):
        return _NA


def _multiplier_pct(value: Any) -> str:
    """레버 delta (0~1) → '+NN.N%'."""
    if value is None:
        return _NA
    try:
        return f"+{float(value) * 100:.1f}%"
    except (TypeError, ValueError):
        return _NA


def _probability_fill(prob: Any) -> PatternFill | None:
    """달성 확률 색상: ≥70% 녹색, 50~69% 노랑, <50% 빨강."""
    if prob is None:
        return None
    try:
        p = float(prob)
    except (TypeError, ValueError):
        return None
    if p >= 70:
        return GREEN_FILL
    if p >= 50:
        return YELLOW_FILL
    return RED_FILL


def _case_label(target_case: str | None) -> str:
    labels = {
        "LEVER": "4-레버 곱셈 산정 (정본)",
        "A": "A 케이스 (성장 잠재력 큼)",
        "B": "B 케이스 (보통)",
        "C": "C 케이스 (방어형)",
        "D": "D 케이스 (운영 정비)",
        "E": "E 케이스 (신규·소형)",
        "manual": "수동 입력",
    }
    if target_case is None:
        return _NA
    return labels.get(target_case, target_case)


# ─────────────────────────────────────────────
# 공용 셀 쓰기 헬퍼
# ─────────────────────────────────────────────
def _set(
    ws: Worksheet,
    coord: str,
    value: Any,
    *,
    font: Font | None = None,
    fill: PatternFill | None = None,
    alignment: Alignment | None = None,
    border: Border | None = THIN_BORDER,
    number_format: str | None = None,
) -> None:
    cell = ws[coord]
    cell.value = value
    cell.font = font or BODY_FONT
    if fill is not None:
        cell.fill = fill
    cell.alignment = alignment or LEFT
    if border is not None:
        cell.border = border
    if number_format is not None:
        cell.number_format = number_format


def _merge_title(
    ws: Worksheet,
    range_: str,
    text: str,
    *,
    font: Font = TITLE_FONT,
    fill: PatternFill = TITLE_FILL,
    height: int = 28,
) -> None:
    ws.merge_cells(range_)
    start = range_.split(":")[0]
    cell = ws[start]
    cell.value = text
    cell.font = font
    cell.fill = fill
    cell.alignment = CENTER
    # 머지 범위의 모든 셀에 fill/테두리 적용
    for row in ws[range_]:
        for c in row:
            c.fill = fill
            c.border = THIN_BORDER
    # 첫 행 높이 — 범위 첫 행 기준
    first_row = int("".join(ch for ch in start if ch.isdigit()))
    ws.row_dimensions[first_row].height = height


def _set_col_widths(ws: Worksheet, widths: dict[str, int]) -> None:
    for col, w in widths.items():
        ws.column_dimensions[col].width = w


# ─────────────────────────────────────────────
# 시트 1 — 요약
# ─────────────────────────────────────────────
def _render_sheet_summary(
    ws: Worksheet,
    plan: SolutionPlan,
    lr: dict[str, Any] | None,
) -> None:
    ws.title = "1.요약"
    _set_col_widths(ws, {"A": 22, "B": 20, "C": 20, "D": 20, "E": 20})

    # 제목 (A1:E1)
    _merge_title(ws, "A1:E1", "목표매출 산정 근거 — 담당자 내부용")

    # ── 매장 메타 (A3~A8) ──
    # 담당자 피드백 (2026-04-26): "업종" 단일 필드면 원 업종과 벤치마크 구분이 모호.
    # "원 업종" + "벤치마크 카테고리" 두 행으로 분리 표시 (두 값이 다를 수 있음).
    store = plan.store
    target_meta = plan.target_meta
    target_case = target_meta.target_case if target_meta else None
    bench = (lr or {}).get("analysis", {}).get("cuisine_benchmark") if lr else None
    orig_type = store.business_type or _NA
    bench_display = bench or _NA
    # 원 업종과 벤치마크가 다르면 화살표 병기로 한 눈에 매핑 확인
    if orig_type != _NA and bench_display != _NA and orig_type != bench_display:
        bench_display = f"{bench_display}  (← 매핑)"

    meta_rows = [
        ("매장명", store.name),
        ("원 업종", orig_type),
        ("벤치마크 카테고리", bench_display),
        ("위치", store.location or _NA),
        ("산정일", store.document_date or _NA),
        ("케이스", f"{target_case or _NA} — {_case_label(target_case)}"),
    ]
    for idx, (label, value) in enumerate(meta_rows, start=3):
        _set(ws, f"A{idx}", label, font=META_LABEL_FONT, fill=ALT_FILL)
        ws.merge_cells(f"B{idx}:E{idx}")
        _set(ws, f"B{idx}", value, font=BODY_FONT)
        # 머지된 셀들도 테두리
        for col in ("C", "D", "E"):
            ws[f"{col}{idx}"].border = THIN_BORDER

    # ── 1차 목표 블록 (A10:E15) ──
    targets = (lr or {}).get("targets") or {}
    t1_raw = targets.get("tier_1_revenue_won")
    t2_raw = targets.get("tier_2_revenue_won")

    _merge_title(ws, "A10:E10", "1차 목표 (3개월, 3% 구간)", font=SUBTITLE_FONT, fill=SUBTITLE_FILL, height=22)

    headers_t = ["항목", "원본 수치", "반올림(업주 표시)", "증가율", "달성 확률"]
    for i, h in enumerate(headers_t):
        _set(ws, f"{get_column_letter(i+1)}11", h, font=BODY_BOLD, fill=HEADER_FILL, alignment=CENTER)
        ws[f"{get_column_letter(i+1)}11"].font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")

    if t1_raw is not None:
        _set(ws, "A12", "1차 목표 매출", font=META_LABEL_FONT)
        _set(ws, "B12", int(t1_raw), alignment=RIGHT, number_format=WON_FORMAT)
        _set(ws, "C12", _round_down_to_500k(int(t1_raw)), alignment=RIGHT, number_format=WON_FORMAT, font=BODY_BOLD)
        _set(ws, "D12", _pct_str(targets.get("tier_1_growth_pct")), alignment=CENTER)
        prob_1 = targets.get("tier_1_probability_pct")
        _set(
            ws,
            "E12",
            f"{prob_1}%" if prob_1 is not None else _NA,
            alignment=CENTER,
            fill=_probability_fill(prob_1),
            font=BODY_BOLD,
        )
        _set(ws, "A13", "예상 월 수수료", font=META_LABEL_FONT, fill=ALT_FILL)
        fee_1 = targets.get("tier_1_monthly_fee_won")
        ws.merge_cells("B13:E13")
        if fee_1 is not None:
            _set(ws, "B13", int(fee_1), alignment=RIGHT, number_format=WON_FORMAT)
        else:
            _set(ws, "B13", _NA, alignment=LEFT)
        for col in ("C", "D", "E"):
            ws[f"{col}13"].border = THIN_BORDER
    else:
        _set(ws, "A12", "1차 목표", font=META_LABEL_FONT)
        ws.merge_cells("B12:E12")
        _set(ws, "B12", "lever_report 미존재 — 산정 불가", font=BODY_FONT)
        for col in ("C", "D", "E"):
            ws[f"{col}12"].border = THIN_BORDER

    # ── 2차 목표 블록 (A17:E23) ──
    _merge_title(ws, "A17:E17", "2차 목표 (6개월, 5% 구간)", font=SUBTITLE_FONT, fill=SUBTITLE_FILL, height=22)
    for i, h in enumerate(headers_t):
        cell = ws[f"{get_column_letter(i+1)}18"]
        cell.value = h
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    if t2_raw is not None:
        _set(ws, "A19", "2차 목표 매출", font=META_LABEL_FONT)
        _set(ws, "B19", int(t2_raw), alignment=RIGHT, number_format=WON_FORMAT)
        _set(ws, "C19", _round_down_to_500k(int(t2_raw)), alignment=RIGHT, number_format=WON_FORMAT, font=BODY_BOLD)
        _set(ws, "D19", _pct_str(targets.get("tier_2_growth_pct")), alignment=CENTER)
        prob_2 = targets.get("tier_2_probability_pct")
        _set(
            ws,
            "E19",
            f"{prob_2}%" if prob_2 is not None else _NA,
            alignment=CENTER,
            fill=_probability_fill(prob_2),
            font=BODY_BOLD,
        )
        _set(ws, "A20", "예상 월 수수료", font=META_LABEL_FONT, fill=ALT_FILL)
        fee_2 = targets.get("tier_2_monthly_fee_won")
        ws.merge_cells("B20:E20")
        if fee_2 is not None:
            _set(ws, "B20", int(fee_2), alignment=RIGHT, number_format=WON_FORMAT)
        else:
            _set(ws, "B20", _NA, alignment=LEFT)
        for col in ("C", "D", "E"):
            ws[f"{col}20"].border = THIN_BORDER

        # 수수료 상한 체크
        fee_cap_ok = targets.get("fee_cap_ok")
        _set(ws, "A21", "수수료 상한 200만원", font=META_LABEL_FONT, fill=ALT_FILL)
        ws.merge_cells("B21:E21")
        if fee_cap_ok is True:
            merge_fill_21: PatternFill | None = GREEN_FILL
            _set(ws, "B21", "적합", font=GREEN_FONT, fill=GREEN_FILL, alignment=CENTER)
        elif fee_cap_ok is False:
            merge_fill_21 = RED_FILL
            _set(ws, "B21", "재조정 필요", font=RED_FONT, fill=RED_FILL, alignment=CENTER)
        else:
            merge_fill_21 = None
            _set(ws, "B21", _NA, alignment=LEFT)
        for col in ("C", "D", "E"):
            ws[f"{col}21"].border = THIN_BORDER
            if merge_fill_21 is not None:
                ws[f"{col}21"].fill = merge_fill_21

        # 조정 사유
        adj_note = targets.get("adjustment_note")
        if adj_note:
            _set(ws, "A22", "조정 사유", font=META_LABEL_FONT, fill=ALT_FILL)
            ws.merge_cells("B22:E22")
            _set(ws, "B22", str(adj_note))
            for col in ("C", "D", "E"):
                ws[f"{col}22"].border = THIN_BORDER
    else:
        _set(ws, "A19", "2차 목표", font=META_LABEL_FONT)
        ws.merge_cells("B19:E19")
        _set(ws, "B19", "lever_report 미존재 — 산정 불가", font=BODY_FONT)
        for col in ("C", "D", "E"):
            ws[f"{col}19"].border = THIN_BORDER

    # ── 플랫폼별 매출 (A25:E29) ──
    _merge_title(ws, "A25:E25", "플랫폼별 매출 (배민 기준)", font=SUBTITLE_FONT, fill=SUBTITLE_FILL, height=22)
    platform_headers = ["플랫폼", "현재", "1차 목표", "2차 목표", "상태"]
    for i, h in enumerate(platform_headers):
        cell = ws[f"{get_column_letter(i+1)}26"]
        cell.value = h
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    # 현재 매출은 core_metrics[0] 값 파싱 시도 (baseline)
    current_rev: int | None = None
    for m in plan.core_metrics:
        if m.is_baseline:
            # "1,800,000원" 같은 문자열에서 숫자만 추출
            digits = "".join(ch for ch in m.value if ch.isdigit())
            if digits:
                current_rev = int(digits)
            break

    t1_display = _round_down_to_500k(int(t1_raw)) if t1_raw is not None else None
    t2_display = _round_down_to_500k(int(t2_raw)) if t2_raw is not None else None

    platforms = [
        ("배민", current_rev, t1_display, t2_display, "주 수집"),
        ("쿠팡이츠", None, None, None, "데이터 미수집"),
        ("요기요", None, None, None, "데이터 미수집"),
    ]
    for idx, (name, cur, t1_p, t2_p, status) in enumerate(platforms, start=27):
        _set(ws, f"A{idx}", name, font=META_LABEL_FONT, alignment=CENTER, fill=ALT_FILL if idx % 2 else None)
        if cur is not None:
            _set(ws, f"B{idx}", cur, alignment=RIGHT, number_format=WON_FORMAT)
        else:
            _set(ws, f"B{idx}", _NA, alignment=CENTER)
        if t1_p is not None:
            _set(ws, f"C{idx}", t1_p, alignment=RIGHT, number_format=WON_FORMAT)
        else:
            _set(ws, f"C{idx}", _NA, alignment=CENTER)
        if t2_p is not None:
            _set(ws, f"D{idx}", t2_p, alignment=RIGHT, number_format=WON_FORMAT)
        else:
            _set(ws, f"D{idx}", _NA, alignment=CENTER)
        status_fill = GREEN_FILL if status == "주 수집" else GREY_FILL
        _set(ws, f"E{idx}", status, alignment=CENTER, fill=status_fill)


# ─────────────────────────────────────────────
# 시트 2 — 레버 분석
# ─────────────────────────────────────────────
def _lever_delta_fill(value: Any) -> tuple[PatternFill | None, Font]:
    """레버 개선폭 색상: 0% 회색, +30%↑ 녹색 볼드, 그 외 기본."""
    if value is None:
        return GREY_FILL, BODY_FONT
    try:
        v = float(value)
    except (TypeError, ValueError):
        return GREY_FILL, BODY_FONT
    if v == 0:
        return GREY_FILL, BODY_FONT
    if v >= 0.30:
        return GREEN_FILL, GREEN_FONT
    return None, BODY_FONT


def _render_sheet_levers(ws: Worksheet, lr: dict[str, Any] | None) -> None:
    ws.title = "2.레버분석"
    _set_col_widths(ws, {"A": 18, "B": 16, "C": 22, "D": 10, "E": 14, "F": 14})

    _merge_title(ws, "A1:F1", "레버별 현황 및 개선 여력")

    if lr is None:
        ws.merge_cells("A3:F3")
        _set(ws, "A3", "lever_report 미존재 — 레버 경로 아님 (E 가드 또는 수동 케이스)", font=BODY_BOLD, alignment=CENTER, fill=GREY_FILL)
        for col in ("B", "C", "D", "E", "F"):
            ws[f"{col}3"].border = THIN_BORDER
            ws[f"{col}3"].fill = GREY_FILL
        return

    analysis = lr.get("analysis") or {}
    current_impressions = lr.get("current_impressions_31d")
    bench_name = analysis.get("cuisine_benchmark") or _NA

    # 담당자 피드백 (2026-04-26): 벤치마크 출처 각주를 헤더 상단에 명시.
    ws.merge_cells("A2:F2")
    _set(
        ws,
        "A2",
        f"* 벤치마크 카테고리: {bench_name}  (매장 '원 업종' → '벤치마크' 매핑 결과)",
        font=Font(name=FONT_NAME, size=9, italic=True, color="595959"),
        fill=ALT_FILL,
        alignment=LEFT,
    )
    for col in ("B", "C", "D", "E", "F"):
        ws[f"{col}2"].border = THIN_BORDER
        ws[f"{col}2"].fill = ALT_FILL

    headers = ["레버", "현재값", "벤치마크", "위치", "단기 개선", "중기 개선"]
    for i, h in enumerate(headers):
        cell = ws[f"{get_column_letter(i+1)}3"]
        cell.value = h
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    imp = analysis.get("impression_delta") or {}
    ctr = analysis.get("ctr_delta") or {}
    cvr = analysis.get("cvr_delta") or {}
    aov = analysis.get("aov_delta") or {}

    # 레버 정의 — (이름, 현재값, 벤치마크 문자열, 위치/비고, short_pct, mid_pct, basis)
    cur_ctr = analysis.get("current_ctr_pct")
    cur_cvr = analysis.get("current_cvr_pct")
    cur_aov = analysis.get("current_aov_won")

    lever_defs = [
        (
            "레버 1 — 노출수",
            f"{current_impressions:,}건" if isinstance(current_impressions, int) else _NA,
            bench_name,
            "-",
            imp.get("short_term_pct"),
            imp.get("mid_term_pct"),
            imp.get("basis"),
        ),
        (
            "레버 2 — CTR",
            _pct_str(cur_ctr, 2),
            f"{bench_name} 카테고리",
            "중간",
            ctr.get("short_term_pct"),
            ctr.get("mid_term_pct"),
            ctr.get("basis"),
        ),
        (
            "레버 3 — CVR",
            _pct_str(cur_cvr, 2),
            f"{bench_name} 카테고리",
            "중간",
            cvr.get("short_term_pct"),
            cvr.get("mid_term_pct"),
            cvr.get("basis"),
        ),
        (
            "레버 4 — 객단가(AOV)",
            f"{int(cur_aov):,}원" if isinstance(cur_aov, int) else _NA,
            f"{bench_name} 탄력성",
            "중간",
            aov.get("short_term_pct"),
            aov.get("mid_term_pct"),
            aov.get("basis"),
        ),
    ]

    row = 4
    for name, cur, bench, pos, short_p, mid_p, basis in lever_defs:
        _set(ws, f"A{row}", name, font=META_LABEL_FONT, fill=ALT_FILL)
        _set(ws, f"B{row}", cur, alignment=CENTER)
        _set(ws, f"C{row}", bench, alignment=CENTER)
        _set(ws, f"D{row}", pos, alignment=CENTER)
        fill_s, font_s = _lever_delta_fill(short_p)
        fill_m, font_m = _lever_delta_fill(mid_p)
        _set(ws, f"E{row}", _multiplier_pct(short_p), alignment=CENTER, fill=fill_s, font=font_s)
        _set(ws, f"F{row}", _multiplier_pct(mid_p), alignment=CENTER, fill=fill_m, font=font_m)
        row += 1
        # 근거 행 (A~F 병합)
        ws.merge_cells(f"A{row}:F{row}")
        basis_text = f"근거: {basis}" if basis else "근거: 제공되지 않음"
        _set(ws, f"A{row}", basis_text, font=BODY_FONT, fill=ALT_FILL, alignment=LEFT)
        for col in ("B", "C", "D", "E", "F"):
            ws[f"{col}{row}"].border = THIN_BORDER
            ws[f"{col}{row}"].fill = ALT_FILL
        row += 1


# ─────────────────────────────────────────────
# 시트 3 — 가정·한계
# ─────────────────────────────────────────────
def _render_sheet_assumptions(
    ws: Worksheet,
    plan: SolutionPlan,
    lr: dict[str, Any] | None,
) -> None:
    ws.title = "3.가정·한계"
    _set_col_widths(ws, {"A": 8, "B": 70})

    _merge_title(ws, "A1:B1", "산정에 사용된 가정")

    analysis = (lr or {}).get("analysis") or {}
    bench = analysis.get("cuisine_benchmark")
    ctr = analysis.get("current_ctr_pct")
    cvr = analysis.get("current_cvr_pct")
    aov = analysis.get("current_aov_won")

    aov_str = f"{int(aov):,}원" if isinstance(aov, int) else _NA

    tam_meta = plan.tam_meta
    if tam_meta is None:
        tam_line = "미연동 (tam_meta 없음)"
    elif tam_meta.available:
        if tam_meta.tam_monthly_revenue_won is not None:
            tam_line = f"연동됨 (월 추정 {int(tam_meta.tam_monthly_revenue_won):,}원)"
        else:
            tam_line = "연동됨"
    else:
        tam_line = f"미연동 — {tam_meta.reason or '사유 미상'}"

    season = (lr or {}).get("season_factors")
    if season:
        season_line = (
            f"tier1 ×{season.get('tier_1', 1.0):.2f} / "
            f"tier2 ×{season.get('tier_2', 1.0):.2f}"
        )
    else:
        season_line = "미반영 (담당자 검토 필요)"

    tam_cap = bool((lr or {}).get("tam_cap_applied"))
    hs = (lr or {}).get("historical_sanity")
    if isinstance(hs, dict) and hs.get("available"):
        hs_line = (
            f"n={hs.get('n')}건 · P50 {hs.get('p50_growth')}% / "
            f"P80 {hs.get('p80_growth')}% / 레버 {hs.get('lever_growth')}% "
            f"— {hs.get('verdict', _NA)}"
        )
    elif isinstance(hs, dict):
        hs_line = f"비교 불가 ({hs.get('reason', _NA)})"
    else:
        hs_line = "데이터 없음"

    assumptions = [
        (
            "카테고리 벤치마크",
            f"{bench or _NA} (CTR {_pct_str(ctr, 2)} · "
            f"CVR {_pct_str(cvr, 2)} · AOV {aov_str})",
        ),
        ("상권 경쟁도(TAM)", tam_line),
        ("계절성 반영", season_line),
        ("TAM 캡(25%)", "적용" if tam_cap else "미적용"),
        ("매장주 실행 준수", "주 1회 이상 협조 전제로 확률 산정"),
        ("내부 실적 대비(sanity)", hs_line),
    ]

    row = 3
    for i, (label, desc) in enumerate(assumptions, start=1):
        _set(ws, f"A{row}", i, font=BODY_BOLD, alignment=CENTER, fill=ALT_FILL)
        _set(ws, f"B{row}", f"{label}: {desc}", font=BODY_FONT, alignment=LEFT)
        row += 1

    # 가드 (옵션)
    guard_note = (lr or {}).get("guard_note")
    if guard_note:
        _set(ws, f"A{row}", "!", font=BODY_BOLD, alignment=CENTER, fill=YELLOW_FILL)
        _set(ws, f"B{row}", f"가드: {guard_note}", font=BODY_BOLD, fill=YELLOW_FILL)
        row += 1

    # 공백 2행
    row += 2

    # 데이터 부족 섹션
    _merge_title(ws, f"A{row}:B{row}", "데이터 부족 항목", font=SUBTITLE_FONT, fill=SUBTITLE_FILL, height=22)
    row += 1

    disclaimers = (lr or {}).get("disclaimers") or []
    if disclaimers:
        for d in disclaimers:
            _set(ws, f"A{row}", "·", font=BODY_BOLD, alignment=CENTER)
            _set(ws, f"B{row}", str(d), font=BODY_FONT, alignment=LEFT)
            row += 1
    else:
        _set(ws, f"A{row}", "·", font=BODY_FONT, alignment=CENTER)
        _set(ws, f"B{row}", "추가 주의사항 없음", font=BODY_FONT, alignment=LEFT)


# ─────────────────────────────────────────────
# 시트 4 — REF 검증
# ─────────────────────────────────────────────
def _render_sheet_reference(ws: Worksheet, plan: SolutionPlan) -> None:
    ws.title = "4.REF검증"
    _set_col_widths(ws, {"A": 10, "B": 32, "C": 10, "D": 60})

    _merge_title(ws, "A1:D1", "정본 §필수 준수 자동 검증")

    headers = ["ID", "규칙", "결과", "상세"]
    for i, h in enumerate(headers):
        cell = ws[f"{get_column_letter(i+1)}3"]
        cell.value = h
        cell.font = Font(name=FONT_NAME, size=10, bold=True, color="FFFFFF")
        cell.fill = HEADER_FILL
        cell.alignment = CENTER
        cell.border = THIN_BORDER

    report = check_reference(plan)
    row = 4
    pass_count = 0
    for r in report.results:
        if r.passed:
            mark = "OK"
            fill = GREEN_FILL
            font = GREEN_FONT
            pass_count += 1
        elif r.severity == "error":
            mark = "FAIL"
            fill = RED_FILL
            font = RED_FONT
        else:
            mark = "WARN"
            fill = YELLOW_FILL
            font = YELLOW_FONT
        _set(ws, f"A{row}", r.rule_id, font=BODY_BOLD, alignment=CENTER)
        _set(ws, f"B{row}", r.rule_name, font=BODY_FONT, alignment=LEFT)
        _set(ws, f"C{row}", mark, font=font, fill=fill, alignment=CENTER)
        _set(ws, f"D{row}", r.detail or "", font=BODY_FONT, alignment=LEFT)
        row += 1

    total = len(report.results)
    row += 1
    if report.all_passed:
        summary_fill = GREEN_FILL
        summary_color = "006100"
    else:
        summary_fill = RED_FILL
        summary_color = "9C0006"
    summary_text = (
        f"종합: {pass_count}/{total} PASS"
        if report.all_passed
        else f"종합: {pass_count}/{total} PASS — error 급 실패 존재, 담당자 검토 필요"
    )
    _merge_title(
        ws,
        f"A{row}:D{row}",
        summary_text,
        font=Font(name=FONT_NAME, size=12, bold=True, color=summary_color),
        fill=summary_fill,
        height=24,
    )


# ─────────────────────────────────────────────
# 공개 API
# ─────────────────────────────────────────────
def build_rationale_xlsx(plan: SolutionPlan) -> Workbook:
    """담당자 내부용 산정 근거 Workbook 생성 (4시트).

    lever_report 부재 시 (E 케이스 등) 시트 2는 축약 메시지,
    시트 1의 1·2차 목표 블록은 "산정 불가" 표기, 나머지 시트는 정상 구성.
    """
    target_meta = plan.target_meta
    lr: dict[str, Any] | None = None
    if target_meta is not None and isinstance(target_meta.lever_report, dict):
        lr = target_meta.lever_report

    wb = Workbook()
    ws1 = wb.active
    _render_sheet_summary(ws1, plan, lr)
    ws2 = wb.create_sheet()
    _render_sheet_levers(ws2, lr)
    ws3 = wb.create_sheet()
    _render_sheet_assumptions(ws3, plan, lr)
    ws4 = wb.create_sheet()
    _render_sheet_reference(ws4, plan)
    return wb


def write_rationale_xlsx(plan: SolutionPlan, out_path: str | Path) -> Path:
    """산정 근거 XLSX 를 파일로 저장.

    Args:
        plan: SolutionPlan (검증 통과 모델).
        out_path: 저장 경로 (.xlsx). 부모 디렉토리 자동 생성.

    Returns:
        실제 저장된 Path.
    """
    path = Path(out_path)
    path.parent.mkdir(parents=True, exist_ok=True)
    wb = build_rationale_xlsx(plan)
    wb.save(path)
    return path
