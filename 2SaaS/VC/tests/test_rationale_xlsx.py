"""담당자 내부용 산정 근거 XLSX 생성 테스트.

2026-04-22 담당자 피드백에 따라 기존 txt 버전을 대체.
4시트 구조: 1.요약 / 2.레버분석 / 3.가정·한계 / 4.REF검증.
"""

from __future__ import annotations

import json
import sys
from pathlib import Path

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.generator.rationale_xlsx import (
    build_rationale_xlsx,
    write_rationale_xlsx,
)
from src.schemas.solution import SolutionPlan


SAMPLES = Path(__file__).parent.parent / "data" / "samples"


def _load_plan(name: str) -> SolutionPlan:
    data = json.loads((SAMPLES / name).read_text(encoding="utf-8"))
    return SolutionPlan(**data)


def test_four_sheet_names() -> None:
    """4개 시트명이 정확히 일치해야 한다."""
    plan = _load_plan("더피플버거_sample.json")
    wb = build_rationale_xlsx(plan)
    assert wb.sheetnames == ["1.요약", "2.레버분석", "3.가정·한계", "4.REF검증"]


def test_summary_sheet_store_and_tier_values() -> None:
    """시트1: 매장명·1차/2차 목표의 원본 수치 + 반올림 값이 병기."""
    plan = _load_plan("더피플버거_sample.json")
    wb = build_rationale_xlsx(plan)
    ws = wb["1.요약"]

    # 매장명 행 (A3=라벨, B3=값)
    assert ws["A3"].value == "매장명"
    assert ws["B3"].value == "더피플버거"

    # 원 업종 행 (A4) — 담당자 피드백(2026-04-26)에 따라 벤치마크와 분리 표시
    assert ws["A4"].value == "원 업종"

    # 벤치마크 카테고리 행 (A5) — 라벨 네이밍 변경
    assert ws["A5"].value == "벤치마크 카테고리"
    assert "한식" in str(ws["B5"].value)

    # 1차 목표 블록 — 12행
    lr = plan.target_meta.lever_report  # type: ignore[union-attr]
    t1 = lr["targets"]["tier_1_revenue_won"]
    assert ws["B12"].value == int(t1)  # 원본
    assert ws["C12"].value == (int(t1) // 500_000) * 500_000  # 반올림

    # 2차 목표 블록 — 19행
    t2 = lr["targets"]["tier_2_revenue_won"]
    assert ws["B19"].value == int(t2)
    assert ws["C19"].value == (int(t2) // 500_000) * 500_000

    # 확률 셀 — 색상 fill 적용 확인
    e12 = ws["E12"]
    assert e12.value == "75%"
    # 75%이면 GREEN (C6EFCE)
    assert "C6EFCE" in str(e12.fill.start_color.rgb).upper()


def test_levers_sheet_four_levers_present() -> None:
    """시트2: 4개 레버(노출·CTR·CVR·AOV) 행이 모두 존재."""
    plan = _load_plan("더피플버거_sample.json")
    wb = build_rationale_xlsx(plan)
    ws = wb["2.레버분석"]

    # 제목
    assert ws["A1"].value == "레버별 현황 및 개선 여력"

    # 헤더 3행
    assert ws["A3"].value == "레버"
    assert ws["B3"].value == "현재값"
    assert ws["C3"].value == "벤치마크"
    assert ws["E3"].value == "단기 개선"
    assert ws["F3"].value == "중기 개선"

    # 레버 행 (각 레버 = 데이터 1행 + 근거 1행) → 4행, 6행, 8행, 10행이 레버
    lever_rows = [4, 6, 8, 10]
    names = [ws[f"A{r}"].value for r in lever_rows]
    assert any("노출수" in str(n) for n in names)
    assert any("CTR" in str(n) for n in names)
    assert any("CVR" in str(n) for n in names)
    assert any("객단가" in str(n) or "AOV" in str(n) for n in names)

    # CTR 레버 단기 +15% 확인 (sample의 ctr_delta.short_term_pct=0.15)
    ctr_row = next(
        r for r in lever_rows if "CTR" in str(ws[f"A{r}"].value)
    )
    assert ws[f"E{ctr_row}"].value == "+15.0%"


def test_assumptions_sheet_has_min_five_items() -> None:
    """시트3: 최소 5개 가정 항목이 번호 매겨져 존재."""
    plan = _load_plan("더피플버거_sample.json")
    wb = build_rationale_xlsx(plan)
    ws = wb["3.가정·한계"]

    # 제목
    assert ws["A1"].value == "산정에 사용된 가정"

    # A3부터 번호(1,2,3,...) 열 스캔 — 최소 5개
    numbered_rows = 0
    for r in range(3, 12):
        v = ws[f"A{r}"].value
        if isinstance(v, int) and 1 <= v <= 10:
            numbered_rows += 1
    assert numbered_rows >= 5

    # 특정 키워드 항목 존재 확인
    all_descriptions = " ".join(
        str(ws[f"B{r}"].value or "") for r in range(3, 12)
    )
    assert "카테고리 벤치마크" in all_descriptions
    assert "TAM" in all_descriptions
    assert "매장주 실행 준수" in all_descriptions


def test_reference_sheet_has_all_eight_rules() -> None:
    """시트4: REF-1~REF-8 8행 모두 존재 + 색상 fill 검증."""
    plan = _load_plan("더피플버거_sample.json")
    wb = build_rationale_xlsx(plan)
    ws = wb["4.REF검증"]

    # 헤더 3행
    assert ws["A3"].value == "ID"
    assert ws["B3"].value == "규칙"
    assert ws["C3"].value == "결과"
    assert ws["D3"].value == "상세"

    # REF-1 ~ REF-8 — 4행부터 11행
    rule_ids = [ws[f"A{r}"].value for r in range(4, 12)]
    assert rule_ids == [
        "REF-1",
        "REF-2",
        "REF-3",
        "REF-4",
        "REF-5",
        "REF-6",
        "REF-7",
        "REF-8",
    ]

    # 각 결과 셀에 OK/FAIL/WARN 중 하나, 색상 fill 존재
    valid_colors = {"C6EFCE", "FFEB9C", "FFC7CE"}  # green, yellow, red
    for r in range(4, 12):
        result_cell = ws[f"C{r}"]
        assert result_cell.value in ("OK", "FAIL", "WARN")
        rgb = str(result_cell.fill.start_color.rgb).upper()
        assert any(color in rgb for color in valid_colors), (
            f"REF row {r} color mismatch: {rgb}"
        )


def test_e_case_without_lever_report() -> None:
    """E 케이스 (lever_report=None) → 시트2는 축약 메시지, 나머지는 정상."""
    plan = _load_plan("더피플버거_sample_e_case.json")
    wb = build_rationale_xlsx(plan)
    # 시트 이름은 4개 유지
    assert wb.sheetnames == ["1.요약", "2.레버분석", "3.가정·한계", "4.REF검증"]

    # 시트1: 매장명 확인
    ws1 = wb["1.요약"]
    assert ws1["B3"].value == "신생버거샵"

    # 시트2: 축약 메시지 "lever_report 미존재" 포함
    ws2 = wb["2.레버분석"]
    assert ws2["A3"].value is not None
    assert "lever_report 미존재" in str(ws2["A3"].value)

    # 시트4: REF 8행은 여전히 존재
    ws4 = wb["4.REF검증"]
    rule_ids = [ws4[f"A{r}"].value for r in range(4, 12)]
    assert rule_ids == [f"REF-{i}" for i in range(1, 9)]


def test_write_rationale_xlsx_round_trip(tmp_path: Path) -> None:
    """write_rationale_xlsx 가 실제 .xlsx 파일을 생성하고 재로드로 셀 값 검증."""
    plan = _load_plan("더피플버거_sample.json")
    out = tmp_path / "산정근거.xlsx"
    result = write_rationale_xlsx(plan, out)
    assert result == out
    assert out.exists()
    assert out.stat().st_size > 0

    # 재로드해서 시트·셀 값 재검증
    wb = load_workbook(out)
    assert wb.sheetnames == ["1.요약", "2.레버분석", "3.가정·한계", "4.REF검증"]

    ws1 = wb["1.요약"]
    assert ws1["B3"].value == "더피플버거"
    # 1차 목표 원본 수치 셀
    lr = plan.target_meta.lever_report  # type: ignore[union-attr]
    assert ws1["B12"].value == int(lr["targets"]["tier_1_revenue_won"])
