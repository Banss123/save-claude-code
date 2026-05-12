"""XLSX 빌더 테스트."""

import json
from pathlib import Path
import sys

import pytest
from openpyxl import load_workbook

sys.path.insert(0, str(Path(__file__).parent.parent))

from src.schemas.menu import MenuPlan
from src.generator.xlsx_builder import build_menu_xlsx


SAMPLES = Path(__file__).parent.parent / "data" / "samples"
OUTPUT = Path(__file__).parent.parent / "output"


def _load_menu_plan() -> MenuPlan:
    with open(SAMPLES / "더피플버거_menu_sample.json", encoding="utf-8") as f:
        return MenuPlan(**json.load(f))


def _load_solution(name: str) -> dict:
    with open(SAMPLES / name, encoding="utf-8") as f:
        return json.load(f)


def test_build_sample_xlsx():
    """더피플버거 샘플 데이터로 XLSX 생성 테스트."""
    plan = _load_menu_plan()
    output_path = OUTPUT / f"{plan.current.store_name}_메뉴판_가안.xlsx"
    result = build_menu_xlsx(plan, output_path)

    assert result.exists()
    assert result.stat().st_size > 0

    print(f"[OK] XLSX 생성 성공: {result}")
    print(f"   파일 크기: {result.stat().st_size:,} bytes")
    print(f"   현안: {plan.current.total_groups}그룹, {plan.current.total_menus}메뉴")
    print(f"   가안: {plan.proposed.total_groups}그룹, {plan.proposed.total_menus}메뉴")
    changed = sum(
        1 for g in plan.proposed.groups for item in g.items if item.is_changed
    )
    print(f"   변경 항목: {changed}건")


def test_xlsx_without_solution_plan_has_two_sheets():
    """solution_plan 미전달 시 기존 현안/가안 시트만 존재."""
    plan = _load_menu_plan()
    output_path = OUTPUT / "더피플버거_no_meta.xlsx"
    build_menu_xlsx(plan, output_path)
    wb = load_workbook(output_path)
    assert wb.sheetnames == ["현안", "가안"]


def test_xlsx_with_tier_plan_does_not_add_meta_sheet():
    """담당자 피드책(2026-04-22): 업주 제공 XLSX 에는 산정 근거 시트를 넣지 않는다.

    solution_plan 을 전달해도 XLSX 는 현안/가안 2시트만 유지.
    '목표매출_근거' 시트가 있으면 업주에게 내부 수치가 노출되므로
    별도 텍스트 파일(rationale_text)로 분리됐다.
    """
    plan = _load_menu_plan()
    sol = _load_solution("더피플버거_sample.json")
    assert sol.get("tier_plan") is not None  # 샘플 데이터 sanity

    output_path = OUTPUT / "더피플버거_with_meta.xlsx"
    build_menu_xlsx(plan, output_path, solution_plan=sol)

    wb = load_workbook(output_path)
    # 산정 근거 시트 부재 + 현안/가안만 존재
    assert "목표매출_근거" not in wb.sheetnames
    assert wb.sheetnames == ["현안", "가안"]


def test_xlsx_with_e_case_still_has_only_two_sheets():
    """tier_plan=None (E케이스) 에서도 XLSX 는 현안/가안 2시트만."""
    plan = _load_menu_plan()
    sol = _load_solution("더피플버거_sample_e_case.json")
    assert sol.get("tier_plan") is None  # 샘플 sanity

    output_path = OUTPUT / "더피플버거_e_case_meta.xlsx"
    build_menu_xlsx(plan, output_path, solution_plan=sol)

    wb = load_workbook(output_path)
    assert "목표매출_근거" not in wb.sheetnames
    assert wb.sheetnames == ["현안", "가안"]


if __name__ == "__main__":
    test_build_sample_xlsx()
